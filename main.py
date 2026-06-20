import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import sys
import os
from adm import abrir_admin

# =========================
# CRIA PLANILHA SE NÃO EXISTIR
# =========================

BASE = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
ARQUIVO = os.path.join(BASE, "dados.xlsx")

if not os.path.exists(ARQUIVO):
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Nome"
    ws["B1"] = "Telefone"
    ws["C1"] = "Numero"

    wb.save(ARQUIVO)




# =========================
# FUNÇÃO REGISTRAR
# =========================

def registrar():

    nome = entry_nome.get().strip()
    telefone = entry_telefone.get().strip()
    numero = entry_numero.get().strip()

    # Verifica campos vazios
    if nome == "" or telefone == "" or numero == "":
        messagebox.showerror(
            "Erro",
            "Preencha todos os campos!"
        )
        return

    wb = load_workbook(ARQUIVO)
    ws = wb.active

    # Verifica se o número já foi vendido
    for linha in ws.iter_rows(min_row=2, values_only=True):

        if linha[2] is not None:

            numero_vendido = str(linha[2])

            if numero_vendido == numero:
                messagebox.showerror(
                    "Erro",
                    f"O número {numero} já foi vendido!"
                )
                return

    # Salva na planilha
    ws.append([
        nome,
        telefone,
        numero
    ])

    wb.save(ARQUIVO)

    messagebox.showinfo(
        "Sucesso",
        f"Venda registrada!\n\n"
        f"Nome: {nome}\n"
        f"Telefone: {telefone}\n"
        f"Número: {numero}"
    )

    # Limpa os campos
    entry_nome.delete(0, tk.END)
    entry_telefone.delete(0, tk.END)
    entry_numero.delete(0, tk.END)

# =========================
# JANELA PRINCIPAL
# =========================

janela = tk.Tk()
janela.title("Sistema de Rifa")
janela.geometry("500x350")

# Título
titulo = tk.Label(
    janela,
    text="Sistema de Rifa",
    font=("Arial", 18, "bold")
)
titulo.pack(pady=10)

# Nome
tk.Label(
    janela,
    text="Nome"
).pack()

entry_nome = tk.Entry(
    janela,
    width=40
)
entry_nome.pack(pady=5)

# Telefone
tk.Label(
    janela,
    text="Telefone"
).pack()

entry_telefone = tk.Entry(
    janela,
    width=40
)
entry_telefone.pack(pady=5)

# Número
tk.Label(
    janela,
    text="Número da Rifa"
).pack()

entry_numero = tk.Entry(
    janela,
    width=20
)
entry_numero.pack(pady=5)

# Botão
btn = tk.Button(
    janela,
    text="Registrar Venda",
    command=registrar
)
btn.pack(pady=10)

# Botão Administrador
admin_btn = tk.Button(
    janela,
    text="Administrador",
    command=abrir_admin,
    fg="black",
    bg="orange"
)
admin_btn.pack(pady=10)

janela.mainloop()