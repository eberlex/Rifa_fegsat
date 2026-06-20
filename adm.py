import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import sys
import os

BASE = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
ARQUIVO = os.path.join(BASE, "dados.xlsx")


def limpar_planilha():
    """Apaga todas as vendas da planilha, mantendo cabeçalho."""
    if not messagebox.askyesno("Confirmar", "Apagar todas as vendas?"):
        return

    if not os.path.exists(ARQUIVO):
        messagebox.showinfo("Info", "Arquivo não encontrado.")
        return

    wb = load_workbook(ARQUIVO)
    ws = wb.active

    # Remove todas as linhas a partir da segunda (mantém cabeçalho)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    wb.save(ARQUIVO)
    messagebox.showinfo("Sucesso", "Planilha limpa.")


def abrir_admin():

    senha_janela = tk.Toplevel()
    senha_janela.title("Administrador")
    senha_janela.geometry("300x200")

    tk.Label(
        senha_janela,
        text="Senha do administrador"
    ).pack(pady=10)

    entrada_senha = tk.Entry(
        senha_janela,
        show="*"
    )
    entrada_senha.pack(pady=10)

    def verificar():

        if entrada_senha.get() == "FEGSAT2026":

            admin = tk.Toplevel()
            admin.title("Painel Admin")
            admin.geometry("400x300")

            tk.Button(
                admin,
                text="Apagar Todas as Vendas",
                fg="black",
                bg="red",
                command=limpar_planilha
            ).pack(pady=20)

            senha_janela.destroy()

        else:

            messagebox.showerror(
                "Erro",
                "Senha incorreta."
            )

    tk.Button(
        senha_janela,
        text="Entrar",
        command=verificar
    ).pack(pady=10)
